from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime


wb = Workbook()
wb = load_workbook('./Results.xlsx')
ws = wb.active
topology =  []
degradation = {}
fps_dldt_cpu32 = {}
fps_dldt_cpu8 = {}

for i in range(2, ws.max_row+1):
    if not(ws.cell(row=i, column=1).value in topology):
        topology.append(ws.cell(row=i, column=1).value)
    if ws.cell(row=i, column=2).value == "dldt_cpu32":
        fps_dldt_cpu32[ws.cell(row=i, column=1).value] = float(ws.cell(row=i, column=3).value)
    else:
        fps_dldt_cpu8[ws.cell(row=i, column=1).value] = float(ws.cell(row=i, column=3).value)

for t in topology :
    if not(t in fps_dldt_cpu8):
        degradation[t] = ''
    elif (fps_dldt_cpu8[t] == -7) or (fps_dldt_cpu8[t] == -3) :
        degradation[t] = "Error"
    else:
        degradation[t] = (fps_dldt_cpu8[t] - fps_dldt_cpu32[t]) / fps_dldt_cpu32[t]

redFill = PatternFill(start_color='FF2400',
                   end_color='FF2400',
                   fill_type='solid')

for i in range(2, ws.max_row+1):
    if ws.cell(row=i, column=2).value == 'dldt_cpu8':
        ws.cell(row=i, column=4, value=degradation[ws.cell(row=i, column=1).value])
        if (ws.cell(row=i, column=4).value !='' and ws.cell(row=i, column=4).value !='Error') and float(ws.cell(row=i, column=4).value) < -5:
            ws.cell(row=i, column=4).fill = redFill

now = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
wb.save('degradation_'+str(now)+'.xlsx')




